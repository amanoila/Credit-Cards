from tabulate import tabulate
import re
from openpyxl import *
from datetime import date, datetime


class Transaction:
	def __init__(self, time, amount, details):
		self.date = time
		self.amount = amount
		self.details = details

	def __str__(self):
		return f"Transaction: {self.amount} on {self.date} for {self.details}"


class CreditCard:
	def __init__(self):
		self.name = None
		self.number = None
		self.transactions = None
		self.sheet = None

	def __str__(self):
		return f"Owner name: {self.name}\nIBAN: {self.number}\nCurrent balance: {self.current_amount()}"

	def load_from_excel(self, excel_sheet):
		self.name = excel_sheet["B1"].value
		self.number = excel_sheet['B2'].value
		self.transactions = import_transactions(excel_sheet)
		self.sheet = excel_sheet
		if not self.validate_iban():
			raise ValueError("IBAN invalid!")

	def validate_iban(self):
		pattern = r'^RO' \
				  r'\d{2}'\
				  r'[A-Z]{4}' \
				  r'\w{16}$'
		match = re.match(pattern, self.number)
		return match

	def print_data(self):
		print('Name: ' + self.name + '\nAccount number: ' + self.number)

	def new_transaction(self, transaction):
		self.transactions.append(transaction)
		self.sheet.append([transaction.date, transaction.amount, transaction.details])

	def deposit(self, amount, req_day=datetime.today()):
		self.new_transaction(Transaction(req_day, amount, 'Deposit'))

	def withdrawal(self, amount, req_day=datetime.today()):
		return 'Withdrawal request rejected. Current balance too low. ' \
			if self.current_amount() < amount \
			else self.new_transaction(Transaction(req_day, -amount, 'Deposit'))

	def current_amount(self):
		amount = 0
		for transaction in self.transactions:
			amount = amount + transaction.amount
		return amount

	def statement(self, start_date=datetime(1900, 1, 1), end_date=datetime.today()):
		stat_list = []
		balance = 0
		for trans in self.transactions:
			if start_date < trans.date < end_date:
				stat_list.append([trans.date, trans.amount, trans.details])
				balance += trans.amount
		self.print_data()
		print(f"Statement from {start_date} to {end_date}:")
		print(f'Current balance: {balance}')
		print(tabulate(stat_list, headers=['No.', "Date", "Amount", "Details"], showindex=True))


def import_transactions(ws):
	row = 4
	transactions_list = []
	while ws[f'B{row}'].value is not None:
		tr_date = ws[f'A{row}'].value
		tr_amount = ws[f'B{row}'].value
		tr_details = ws[f'C{row}'].value
		transactions_list.append(Transaction(tr_date, tr_amount, tr_details))
		row += 1
	return transactions_list


cards_list = load_workbook('creditCards.xlsx', data_only=True)

pers1 = CreditCard()
pers1.load_from_excel(cards_list["1"])
print(pers1)

pers2 = CreditCard()
pers2.load_from_excel(cards_list['2'])

pers3 = CreditCard()
pers3.load_from_excel(cards_list['3'])

new_tran = Transaction(datetime(2022, 12, 10), 200, 'Scholarship')
pers1.new_transaction(new_tran)
pers3.withdrawal(250)
pers1.statement()
cards_list.save('output.xlsx')
pers2.statement(datetime(2023, 1, 1), datetime(2023, 3, 15))
